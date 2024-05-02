import pandas as pd
from google.cloud import bigquery
from openpyxl import load_workbook

# 対象データセットの設定
PROJECT_ID = "bigquery-public-data"
DATASET_ID = "chicago_taxi_trips"

# テンプレートファイルの設定
STRT_TBL = (2, 3) # テーブル情報の開始位置 (行, 列)
STRT_COL = (8, 1) # カラム情報の開始位置 (行, 列)

# BigQueryクライアントのインスタンスを生成
client = bigquery.Client()

query = f"""
    SELECT
      * 
    FROM
      `{PROJECT_ID}.{DATASET_ID}.INFORMATION_SCHEMA.COLUMNS`;
"""

# クエリを実行してDataFrameとして取得
df = client.query(query).to_dataframe()
df_columns = df[
    [
        "table_catalog",
        "table_schema",
        "table_name",
        "column_name",
        "ordinal_position",
        "data_type",
        "is_nullable",
        "is_partitioning_column",
        "clustering_ordinal_position",
    ]
]
# Excel書き込み時のNA対策でstr型に変換
df_columns = df_columns.astype({"clustering_ordinal_position": str})

query = f"""
    SELECT
      * 
    FROM
      `{PROJECT_ID}.{DATASET_ID}.INFORMATION_SCHEMA.COLUMN_FIELD_PATHS`;
"""

# クエリを実行してDataFrameとして取得
df = client.query(query).to_dataframe()
df_description = df[["table_name", "column_name", "description"]]

df_table_def = df_columns.merge(df_description, on=["table_name", "column_name"])

# テンプレートファイルを開く
wb = load_workbook("./table_template.xlsx")

# templateシートを取得
ws_template = wb["template"]

# tableごとにテーブル定義書を作成
for table_name, df in df_table_def.groupby("table_name"):

    # templeteシートをコピペ
    ws = wb.copy_worksheet(ws_template)
    print(f"Table Name: {table_name}")

    # シート名をテーブル名に変更
    ws.title = table_name

    # テーブル情報の書き込み
    ws.cell(
        row=STRT_TBL[0],
        column=STRT_TBL[1],
        value=df["table_catalog"].iloc[0],
    )

    ws.cell(
        row=STRT_TBL[0] + 1,
        column=STRT_TBL[1],
        value=df["table_schema"].iloc[0],
    )

    ws.cell(
        row=STRT_TBL[0] + 2,
        column=STRT_TBL[1],
        value=df["table_name"].iloc[0],
    )

    # カラム情報の書き込み
    for i, (_, sr) in enumerate(df.iterrows()):
        row = STRT_COL[0] + i

        ws.cell(row=row, column=STRT_COL[1], value=sr["ordinal_position"])
        ws.cell(row=row, column=STRT_COL[1] + 1, value=sr["column_name"])
        ws.cell(row=row, column=STRT_COL[1] + 2, value=sr["data_type"])
        ws.cell(row=row, column=STRT_COL[1] + 3, value=sr["is_nullable"])
        ws.cell(row=row, column=STRT_COL[1] + 4, value=sr["is_partitioning_column"])
        ws.cell(row=row, column=STRT_COL[1] + 5, value=sr["clustering_ordinal_position"])
        ws.cell(row=row, column=STRT_COL[1] + 6, value=sr["description"])

# templeteシートを削除
wb.remove(ws_template)

# テーブル定義書（エクセルファイル）の保存
save_path = f"./table_definition_{PROJECT_ID}.{DATASET_ID}.xlsx"
wb.save(save_path)
print(f"Saved file at {save_path}")